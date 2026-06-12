from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Optional

import openpyxl

# All column indices are 1-based (openpyxl convention)
_COL_PREDICTION        = 4   # D  — Local / Visitante / Empate (group) or team name (knockout)
_COL_KICKOFF           = 24  # X  — match datetime in the user's local timezone
_COL_HOME_TEAM         = 27  # AA — home team name (Spanish)
_COL_PREDICTED_HOME_G  = 29  # AC — user's predicted home goals
_COL_PREDICTED_AWAY_G  = 30  # AD — user's predicted away goals
_COL_AWAY_TEAM         = 32  # AF — away team name (Spanish)
_COL_MATCH_NUMBER      = 34  # AH — sequential match number (1-104)
_COL_ROUND_LABEL       = 26  # Z  — "J1"/"J2"/"J3" (group) or "1/16"/"1/8"/… (knockout)

_GROUP_OUTCOME_MAP = {
    "Local": "home",
    "Visitante": "away",
    "Empate": "draw",
}


@dataclass
class MatchPrediction:
    match_number: int
    home_team: str
    away_team: str
    kickoff_utc: Optional[datetime]
    prediction: str              # "home" / "away" / "draw"  OR  Spanish team name (knockouts)
    predicted_home_goals: Optional[int]
    predicted_away_goals: Optional[int]
    round_label: str


def parse_name(file_bytes: bytes) -> str:
    """Extract the participant's name from the Home sheet, cell C10."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    try:
        ws = wb["Home"]
        value = ws["C10"].value
        return str(value).strip() if value else ""
    except Exception:
        return ""


def parse_utc_offset(file_bytes: bytes) -> float:
    """Read the UTC offset (in hours) from the Home sheet, cell C8.

    Accepts numeric values (2, -3) or strings like '+2', 'UTC+2', '+05:30'.
    Falls back to 0 (UTC) if unparseable.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    try:
        ws = wb["Home"]
        value = ws["C8"].value
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().upper().replace("UTC", "").replace("GMT", "").replace(" ", "")
        if ":" in s:
            sign = -1 if s.startswith("-") else 1
            parts = s.lstrip("+-").split(":")
            return sign * (float(parts[0]) + float(parts[1]) / 60)
        return float(s)
    except Exception:
        return 0.0


def parse_predictions(file_bytes: bytes, utc_offset_hours: int = 2) -> dict[int, MatchPrediction]:
    """
    Parse an Excel Porra Mundial file and return predictions keyed by match number.

    Args:
        file_bytes: Raw .xlsx content.
        utc_offset_hours: UTC offset of the timezone shown in the Excel
                          (default 2 = Spain UTC+2 summer time).
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb["WORLDCUP"]
    offset = timedelta(hours=utc_offset_hours)

    predictions: dict[int, MatchPrediction] = {}

    for row_idx in range(1, (ws.max_row or 200) + 1):
        raw_match_num = ws.cell(row=row_idx, column=_COL_MATCH_NUMBER).value
        if not isinstance(raw_match_num, (int, float)) or raw_match_num <= 0:
            continue
        match_num = int(raw_match_num)

        home_team = ws.cell(row=row_idx, column=_COL_HOME_TEAM).value
        away_team = ws.cell(row=row_idx, column=_COL_AWAY_TEAM).value
        if not home_team or not away_team:
            continue

        raw_pred = ws.cell(row=row_idx, column=_COL_PREDICTION).value
        prediction = _GROUP_OUTCOME_MAP.get(str(raw_pred), str(raw_pred) if raw_pred else "")

        kickoff_utc = None
        raw_kickoff = ws.cell(row=row_idx, column=_COL_KICKOFF).value
        if isinstance(raw_kickoff, datetime):
            local_dt = raw_kickoff.replace(tzinfo=None)
            kickoff_utc = (local_dt - offset).replace(tzinfo=timezone.utc)

        def _int_or_none(val) -> Optional[int]:
            try:
                return int(val) if val is not None else None
            except (TypeError, ValueError):
                return None

        round_label = ws.cell(row=row_idx, column=_COL_ROUND_LABEL).value or ""

        predictions[match_num] = MatchPrediction(
            match_number=match_num,
            home_team=str(home_team),
            away_team=str(away_team),
            kickoff_utc=kickoff_utc,
            prediction=prediction,
            predicted_home_goals=_int_or_none(ws.cell(row=row_idx, column=_COL_PREDICTED_HOME_G).value),
            predicted_away_goals=_int_or_none(ws.cell(row=row_idx, column=_COL_PREDICTED_AWAY_G).value),
            round_label=str(round_label),
        )

    return predictions


def to_json(predictions: dict[int, MatchPrediction]) -> dict:
    """Serialize predictions to a JSON-safe dict for database storage."""
    return {
        str(k): {
            "home_team": v.home_team,
            "away_team": v.away_team,
            "kickoff_utc": v.kickoff_utc.isoformat() if v.kickoff_utc else None,
            "prediction": v.prediction,
            "predicted_home_goals": v.predicted_home_goals,
            "predicted_away_goals": v.predicted_away_goals,
            "round_label": v.round_label,
        }
        for k, v in predictions.items()
    }
